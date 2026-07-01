"""
Usługi analityki śledczej dla audytu finansowego.
"""

from __future__ import annotations

from collections import Counter, defaultdict
import math
import re

from financial_audit import FormulaParser


def _serialize_excel_value(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def _value_to_text(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value).strip()


def _coerce_number(value):
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _value_to_text(value)
    if not text:
        return None
    text = text.replace(" ", "").replace("\xa0", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _header_kind(header: str) -> str:
    h = (header or "").strip().lower()
    if re.search(r"(faktur|invoice|dokument|numer|nr|rachun|pozycj|id)", h):
        return "document"
    if re.search(r"(kontrah|vendor|dostawc|odbior|klient|firma|nazwa)", h):
        return "party"
    if re.search(r"(kwot|amount|wart|netto|brutto|saldo|koszt|suma|payment|przelew)", h):
        return "amount"
    if re.search(r"(data|date|godzin|czas)", h):
        return "datetime"
    return "generic"


def _leading_digit(value) -> int | None:
    number = _coerce_number(value)
    if number is None:
        return None
    number = abs(number)
    if number < 1:
        return None
    while number >= 10:
        number /= 10
    return int(str(number)[0])


def _benford_expected() -> dict[int, float]:
    return {digit: math.log10(1 + 1 / digit) for digit in range(1, 10)}


def _normalize_formula_pattern(formula: str, current_row: int, current_col: int) -> str:
    from openpyxl.utils import column_index_from_string

    def repl(match):
        sheet = (match.group("sheet") or "").replace("$", "")
        col = match.group("col").replace("$", "")
        row = int(match.group("row"))
        col_offset = column_index_from_string(col) - current_col
        row_offset = row - current_row
        prefix = f"{sheet}!" if sheet else ""
        return f"{prefix}C[{col_offset}]R[{row_offset}]"

    expr = (formula or "").strip().upper().replace(";", ",")
    expr = re.sub(
        r"(?:(?P<sheet>'[^']+'|[A-Z0-9_]+)!)?\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)",
        repl,
        expr,
    )
    return re.sub(r"\s+", "", expr)


def _split_excel_args(args_text: str) -> list[str]:
    args: list[str] = []
    current: list[str] = []
    depth = 0
    in_string = False

    for ch in args_text or "":
        if ch == '"':
            in_string = not in_string
            current.append(ch)
        elif not in_string and ch == "(":
            depth += 1
            current.append(ch)
        elif not in_string and ch == ")":
            depth = max(0, depth - 1)
            current.append(ch)
        elif not in_string and depth == 0 and ch in {",", ";"}:
            args.append("".join(current).strip())
            current = []
        else:
            current.append(ch)

    tail = "".join(current).strip()
    if tail:
        args.append(tail)
    return args


def _parse_excel_coord(cell_ref: str) -> tuple[int | None, int | None]:
    from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

    coord = (cell_ref or "").strip()
    if not coord:
        return None, None
    coord = coord.split("!")[-1].strip()
    try:
        col, row = coordinate_from_string(coord)
        return int(row), int(column_index_from_string(col))
    except Exception:
        return None, None


def _fmt_decimal(value: float) -> str:
    text = f"{value:.6f}".rstrip("0").rstrip(".")
    return text or "0"


def _get_formula_display_value(values_ws, row_idx: int, col_idx: int):
    if values_ws is None:
        return None
    return values_ws.cell(row_idx, col_idx).value


def _append_finding(
    findings: list,
    appendix: list,
    *,
    finding_type: str,
    severity: str,
    sheet: str,
    message: str,
    cell: str = "",
    value=None,
    formula: str = "",
    comment: str = "",
    details: dict | None = None,
):
    finding = {
        "type": finding_type,
        "severity": severity,
        "sheet": sheet,
        "cell": cell,
        "message": message,
        "value": _serialize_excel_value(value),
        "formula": formula or None,
        "comment": comment or message,
    }
    if details:
        finding.update(details)
    findings.append(finding)
    appendix.append(
        {
            "sheet": sheet,
            "cell": cell or "—",
            "value": _serialize_excel_value(value),
            "formula": formula or "",
            "severity": severity,
            "type": finding_type,
            "comment": comment or message,
        }
    )


def _sheet_column_profiles(
    ws, values_ws, selected_rows: set[int] | None = None, selected_cols: set[int] | None = None
):
    profiles = []
    narrowed_rows = sorted(row for row in (selected_rows or set()) if 2 <= row <= ws.max_row)
    full_row_range = range(2, ws.max_row + 1)
    for col_idx in range(1, ws.max_column + 1):
        if selected_cols is not None and selected_rows is None and col_idx not in selected_cols:
            continue
        header = _value_to_text(ws.cell(1, col_idx).value) or f"Kolumna {col_idx}"
        entries = []
        if selected_rows is None:
            row_iter = full_row_range
        elif selected_cols is not None and col_idx in selected_cols:
            row_iter = full_row_range
        else:
            row_iter = narrowed_rows
            if not row_iter:
                continue

        for row_idx in row_iter:
            cell = ws.cell(row_idx, col_idx)
            raw_value = cell.value
            is_formula = isinstance(raw_value, str) and raw_value.startswith("=")
            display_value = _get_formula_display_value(values_ws, row_idx, col_idx) if is_formula else raw_value
            if raw_value in (None, "") and display_value in (None, ""):
                continue
            entries.append(
                {
                    "row": row_idx,
                    "coord": cell.coordinate,
                    "formula": raw_value if is_formula else "",
                    "is_formula": is_formula,
                    "display_value": display_value,
                    "raw_value": raw_value,
                }
            )
        if (
            not entries
            and selected_rows is not None
            and selected_cols is not None
            and col_idx not in selected_cols
        ):
            continue
        profiles.append(
            {
                "col_idx": col_idx,
                "header": header,
                "kind": _header_kind(header),
                "entries": entries,
            }
        )
    return profiles


def _iter_formula_entries(ws, values_ws, *, selected_rows: set[int] | None = None, selected_cols: set[int] | None = None):
    for cell in ws._cells.values():
        raw_value = cell.value
        if not (isinstance(raw_value, str) and raw_value.startswith("=")):
            continue
        row_idx = cell.row
        col_idx = cell.column
        if selected_rows is not None and row_idx not in selected_rows:
            continue
        if selected_cols is not None and col_idx not in selected_cols:
            continue
        yield {
            "coord": cell.coordinate,
            "row": row_idx,
            "col_idx": col_idx,
            "formula": raw_value,
            "display_value": _get_formula_display_value(values_ws, row_idx, col_idx),
        }


def _resolve_sum_like_references(args_text: str, values_ws, current_sheet: str):
    from openpyxl.utils.cell import get_column_letter, range_boundaries

    coords: list[str] = []
    total = 0.0
    had_numeric = False

    for token in _split_excel_args(args_text):
        cleaned = token.strip()
        if not cleaned or "(" in cleaned:
            return None
        if "!" in cleaned:
            sheet_part, ref_part = cleaned.split("!", 1)
            ref_sheet = sheet_part.strip("'")
            if ref_sheet != current_sheet:
                return None
            cleaned = ref_part
        if ":" in cleaned:
            try:
                min_col, min_row, max_col, max_row = range_boundaries(cleaned)
            except Exception:
                return None
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    value = values_ws.cell(row_idx, col_idx).value
                    number = _coerce_number(value)
                    if number is None:
                        continue
                    had_numeric = True
                    total += number
                    coords.append(f"{get_column_letter(col_idx)}{row_idx}")
            continue
        row_idx, col_idx = _parse_excel_coord(cleaned)
        if row_idx is None or col_idx is None:
            return None
        value = values_ws.cell(row_idx, col_idx).value
        number = _coerce_number(value)
        if number is None:
            continue
        had_numeric = True
        total += number
        coords.append(cleaned.replace("$", ""))

    if not had_numeric:
        return None
    return total, coords


def detect_hardcoded_values_and_pattern_deviations(
    sheet_name: str,
    ws,
    values_ws,
    findings: list,
    appendix: list,
    *,
    selected_rows: set[int] | None = None,
    selected_cols: set[int] | None = None,
):
    from openpyxl.utils import get_column_letter

    column_profiles = _sheet_column_profiles(
        ws, values_ws, selected_rows=selected_rows, selected_cols=selected_cols
    )
    formula_counts_by_row = Counter(
        entry["row"]
        for profile in column_profiles
        for entry in profile["entries"]
        if entry["is_formula"]
    )
    summary = {
        "hardcoded_values_in_formula_regions": [],
        "formula_pattern_deviations": [],
        "hardcoded_output_like_values": [],
    }

    for profile in column_profiles:
        entries = profile["entries"]
        formula_entries = [entry for entry in entries if entry["is_formula"]]
        if len(formula_entries) < 2:
            continue

        formula_rows = sorted(entry["row"] for entry in formula_entries)
        min_row = formula_rows[0]
        max_row = formula_rows[-1]
        region_entries = [entry for entry in entries if min_row <= entry["row"] <= max_row]
        formula_ratio = len(formula_entries) / max(1, len(region_entries))

        patterns = Counter(
            _normalize_formula_pattern(entry["formula"], entry["row"], profile["col_idx"])
            for entry in formula_entries
        )
        dominant_pattern, dominant_count = patterns.most_common(1)[0]

        if len(formula_entries) >= 3 and dominant_count >= 2:
            for entry in formula_entries:
                normalized = _normalize_formula_pattern(entry["formula"], entry["row"], profile["col_idx"])
                if normalized != dominant_pattern:
                    _append_finding(
                        findings,
                        appendix,
                        finding_type="formula_pattern_deviation",
                        severity="MEDIUM",
                        sheet=sheet_name,
                        cell=entry["coord"],
                        value=entry["display_value"],
                        formula=entry["formula"],
                        message=(
                            f"Formuła w {entry['coord']} odbiega od dominującego wzorca kolumny "
                            f"„{profile['header']}”."
                        ),
                        comment="Sprawdź odwołania względem sąsiednich wierszy; to może być ręczna zmiana logiki.",
                        details={"header": profile["header"]},
                    )
                    summary["formula_pattern_deviations"].append(
                        {
                            "cell": entry["coord"],
                            "header": profile["header"],
                            "formula": entry["formula"],
                            "dominant_pattern": dominant_pattern,
                        }
                    )

        if formula_ratio >= 0.6:
            for entry in region_entries:
                if entry["is_formula"] or entry["display_value"] in (None, ""):
                    continue
                has_formula_above = any(row < entry["row"] for row in formula_rows)
                has_formula_below = any(row > entry["row"] for row in formula_rows)
                if not (has_formula_above and has_formula_below):
                    continue
                _append_finding(
                    findings,
                    appendix,
                    finding_type="hardcoded_in_formula_region",
                    severity="HIGH",
                    sheet=sheet_name,
                    cell=entry["coord"],
                    value=entry["display_value"],
                    message=(
                        f"W obszarze formuł kolumny „{profile['header']}” wykryto twardą wartość w {entry['coord']}."
                    ),
                    comment="Większość komórek w tej części kolumny zawiera formuły, więc wpis ręczny wymaga wyjaśnienia.",
                    details={"header": profile["header"], "column": get_column_letter(profile["col_idx"])},
                )
                summary["hardcoded_values_in_formula_regions"].append(
                    {
                        "cell": entry["coord"],
                        "header": profile["header"],
                        "value": _serialize_excel_value(entry["display_value"]),
                    }
                )

        for entry in entries:
            if entry["is_formula"] or entry["display_value"] in (None, ""):
                continue
            if entry["row"] <= max_row or entry["row"] > max_row + 2:
                continue
            if formula_counts_by_row.get(entry["row"], 0) == 0:
                continue
            _append_finding(
                findings,
                appendix,
                finding_type="hardcoded_output_like_value",
                severity="HIGH",
                sheet=sheet_name,
                cell=entry["coord"],
                value=entry["display_value"],
                message=(
                    f"Komórka {entry['coord']} wygląda jak wynik końcowy, ale zawiera wpis ręczny "
                    f"zamiast formuły w kolumnie „{profile['header']}”."
                ),
                comment="W tym obszarze wynik powinien być liczony automatycznie, a nie dopisany ręcznie.",
                details={"header": profile["header"], "row": entry["row"]},
            )
            summary["hardcoded_output_like_values"].append(
                {
                    "cell": entry["coord"],
                    "header": profile["header"],
                    "value": _serialize_excel_value(entry["display_value"]),
                }
            )

    return summary, column_profiles


def detect_duplicate_and_numeric_signals(sheet_name: str, column_profiles, findings: list, appendix: list):
    """
    Oblicza wyłącznie sygnały danych o priorytecie HIGH/CRITICAL.

    Celowo pomija słabsze heurystyki (LOW/MEDIUM), których raport nie pokazuje:
    duplikaty kontrahentów i kwot, luki numeracji, kwoty tuż pod progami,
    kwoty okrągłe oraz aktywność nocną/weekendową.
    """
    summary = {
        "duplicates": [],
        "amount_outliers": [],
        "benford_deviations": [],
    }

    for profile in column_profiles:
        entries = profile["entries"]
        kind = profile["kind"]
        header = profile["header"]

        if kind == "document":
            grouped = defaultdict(list)
            for entry in entries:
                if entry["is_formula"]:
                    continue
                norm = _value_to_text(entry["display_value"]).casefold()
                if norm:
                    grouped[norm].append(entry)
            for rows in grouped.values():
                if len(rows) < 2:
                    continue
                row_list = [row["row"] for row in rows]
                _append_finding(
                    findings,
                    appendix,
                    finding_type="duplicate_document",
                    severity="HIGH",
                    sheet=sheet_name,
                    cell=", ".join(row["coord"] for row in rows[:6]),
                    value=rows[0]["display_value"],
                    message=f"W kolumnie „{header}” wykryto powtarzającą się wartość: {_value_to_text(rows[0]['display_value'])}.",
                    comment=f"Wystąpienia w wierszach: {', '.join(map(str, row_list[:12]))}",
                    details={"header": header, "rows": row_list[:25]},
                )
                summary["duplicates"].append(
                    {
                        "header": header,
                        "value": _serialize_excel_value(rows[0]["display_value"]),
                        "rows": row_list[:25],
                        "kind": "document",
                    }
                )

        if kind == "amount":
            numeric_entries = []
            for entry in entries:
                if entry["is_formula"]:
                    continue
                amount = _coerce_number(entry["display_value"])
                if amount is None or amount <= 0:
                    continue
                numeric_entries.append((amount, entry))

            if len(numeric_entries) >= 6:
                values = [amount for amount, _ in numeric_entries]
                mean = sum(values) / len(values)
                variance = sum((value - mean) ** 2 for value in values) / len(values)
                stddev = math.sqrt(variance)
                if stddev > 0:
                    for amount, entry in numeric_entries:
                        z_score = abs((amount - mean) / stddev)
                        if z_score < 4.5:
                            continue
                        _append_finding(
                            findings,
                            appendix,
                            finding_type="amount_outlier",
                            severity="HIGH",
                            sheet=sheet_name,
                            cell=entry["coord"],
                            value=entry["display_value"],
                            message=f"Kwota w {entry['coord']} ekstremalnie odstaje od reszty kolumny „{header}”.",
                            comment=f"Średnia kolumny to {_fmt_decimal(mean)}, odchylenie standardowe {_fmt_decimal(stddev)}, z-score={z_score:.2f}.",
                            details={"header": header, "z_score": round(z_score, 2), "mean": round(mean, 6), "stddev": round(stddev, 6)},
                        )
                        summary["amount_outliers"].append(
                            {
                                "cell": entry["coord"],
                                "header": header,
                                "value": _serialize_excel_value(entry["display_value"]),
                                "z_score": round(z_score, 2),
                            }
                        )

                leading_digits = [digit for digit in (_leading_digit(amount) for amount, _ in numeric_entries) if digit]
                if len(leading_digits) >= 25:
                    expected = _benford_expected()
                    observed_counts = Counter(leading_digits)
                    sample_size = len(leading_digits)
                    observed = {digit: observed_counts.get(digit, 0) / sample_size for digit in range(1, 10)}
                    mad = sum(abs(observed[digit] - expected[digit]) for digit in range(1, 10)) / 9
                    if mad >= 0.05:
                        top_digit = max(range(1, 10), key=lambda digit: abs(observed[digit] - expected[digit]))
                        _append_finding(
                            findings,
                            appendix,
                            finding_type="benford_deviation",
                            severity="HIGH",
                            sheet=sheet_name,
                            message=f"Kolumna „{header}” drastycznie odbiega od oczekiwanego rozkładu cyfr wiodących Benforda.",
                            comment=(
                                f"Największe odchylenie dotyczy cyfry {top_digit}: obserwacja "
                                f"{observed[top_digit] * 100:.1f}% vs oczekiwane {expected[top_digit] * 100:.1f}% (MAD={mad:.3f})."
                            ),
                            details={"header": header, "sample_size": sample_size, "mad": round(mad, 4), "dominant_digit": top_digit},
                        )
                        summary["benford_deviations"].append(
                            {"header": header, "sample_size": sample_size, "mad": round(mad, 4), "dominant_digit": top_digit}
                        )

    return summary


def detect_control_total_and_hidden_reference_signals(
    sheet_name: str,
    ws,
    values_ws,
    findings: list,
    appendix: list,
    *,
    selected_rows: set[int] | None = None,
    selected_cols: set[int] | None = None,
):
    workbook = ws.parent
    summary = {
        "control_total_mismatches": [],
        "cross_sheet_hidden_references": [],
    }
    hidden_ref_seen: set[tuple[str, str]] = set()

    for entry in _iter_formula_entries(
        ws,
        values_ws,
        selected_rows=selected_rows,
        selected_cols=selected_cols,
    ):
        formula = entry["formula"]
        coord = entry["coord"]
        display_value = entry["display_value"]

        for dep in FormulaParser.parse_cell_refs(formula, default_sheet=sheet_name):
            dep_sheet = dep.sheet or sheet_name
            if dep_sheet == sheet_name or dep_sheet not in workbook.sheetnames:
                continue
            dep_state = getattr(workbook[dep_sheet], "sheet_state", "visible") or "visible"
            if dep_state == "visible":
                continue
            fingerprint = (coord, dep_sheet)
            if fingerprint in hidden_ref_seen:
                continue
            hidden_ref_seen.add(fingerprint)
            _append_finding(
                findings,
                appendix,
                finding_type="cross_sheet_hidden_reference",
                severity="HIGH",
                sheet=sheet_name,
                cell=coord,
                value=display_value,
                formula=formula,
                message=f"Formuła w {coord} odwołuje się do ukrytego arkusza „{dep_sheet}”.",
                comment=f"Źródłowy arkusz ma stan {dep_state} i nie jest jawny w zwykłym widoku.",
                details={
                    "source_sheet": dep_sheet,
                    "source_sheet_state": dep_state,
                    "source_cell": f"{dep.col}{dep.row}",
                },
            )
            summary["cross_sheet_hidden_references"].append(
                {
                    "cell": coord,
                    "source_sheet": dep_sheet,
                    "source_sheet_state": dep_state,
                    "source_cell": f"{dep.col}{dep.row}",
                }
            )

        sum_match = re.match(r"^=SUM\((.*)\)$", formula.strip(), re.IGNORECASE)
        if not sum_match:
            continue
        stored_total = _coerce_number(display_value)
        if stored_total is None:
            continue
        resolved = _resolve_sum_like_references(sum_match.group(1), values_ws, sheet_name)
        if not resolved:
            continue
        calculated_total, source_cells = resolved
        diff = abs(stored_total - calculated_total)
        base = max(abs(calculated_total), 1.0)
        if diff <= 0.02 or diff / base <= 0.001:
            continue

        _append_finding(
            findings,
            appendix,
            finding_type="control_total_mismatch",
            severity="HIGH",
            sheet=sheet_name,
            cell=coord,
            value=display_value,
            formula=formula,
            message=f"Suma kontrolna w {coord} nie zgadza się z przeliczeniem zakresu źródłowego.",
            comment=(
                f"Formuła zwraca {_fmt_decimal(stored_total)}, a z danych źródłowych wychodzi "
                f"{_fmt_decimal(calculated_total)}."
            ),
            details={
                "source_cells": source_cells[:30],
                "calculated_total": round(calculated_total, 6),
                "difference": round(diff, 6),
            },
        )
        summary["control_total_mismatches"].append(
            {
                "cell": coord,
                "source_cells": source_cells[:30],
                "calculated_total": round(calculated_total, 6),
                "stored_total": round(stored_total, 6),
            }
        )

    return summary
