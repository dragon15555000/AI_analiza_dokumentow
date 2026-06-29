"""Testy routes finansowych."""

import unittest
import tempfile
from pathlib import Path


class TestFinancialRoutes(unittest.TestCase):
    """Testy API audytu finansowego."""

    def setUp(self):
        """Przygotuj test Excel."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl nie zainstalowany")

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Sheet1"

        # Nagłówki
        self.ws["A1"] = "Price"
        self.ws["B1"] = "Quantity"
        self.ws["C1"] = "Total"

        # Dane z formułami
        self.ws["A2"] = 100
        self.ws["B2"] = 5
        self.ws["C2"] = "=A2*B2"

        self.ws["A3"] = 50
        self.ws["B3"] = 10
        self.ws["C3"] = "=A3*B3"

        self.temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        self.temp_path = Path(self.temp_file.name)
        self.wb.save(self.temp_path)
        self.temp_file.close()

    def tearDown(self):
        """Usuń plik tymczasowy."""
        try:
            self.temp_path.unlink()
        except Exception:
            pass

    def test_extract_formulas_from_excel(self):
        """Test: Ekstraktuj formuły z Excel."""
        from routes.financial_routes import _extract_formulas_from_excel

        result = _extract_formulas_from_excel(self.temp_path)

        self.assertTrue(result["success"])
        self.assertIn("Sheet1", result["sheets"])
        self.assertEqual(len(result["sheets"]["Sheet1"]), 2)

        formulas = result["sheets"]["Sheet1"]
        self.assertEqual(formulas[0]["formula"], "=A2*B2")
        self.assertEqual(formulas[1]["formula"], "=A3*B3")

    def test_financial_audit_via_analyzer(self):
        """Test: Analizuj formuły."""
        from financial_audit import AuditAnalyzer, CellRef

        analyzer = AuditAnalyzer()

        ref_c2 = CellRef("Sheet1", "C", 2)
        analyzer.add_cell(ref_c2, "=A2*B2", value=500)

        ref_c3 = CellRef("Sheet1", "C", 3)
        analyzer.add_cell(ref_c3, "=A3*B3", value=500)

        analyzer.build_dependency_graph()
        analyzer.detect_anomalies()

        summary = analyzer.get_summary()
        self.assertEqual(summary["total_formulas"], 2)
        self.assertEqual(summary["risk_level"], "LOW")


if __name__ == "__main__":
    unittest.main()
