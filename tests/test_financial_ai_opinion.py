import json
from io import BytesIO
from unittest.mock import patch

from openpyxl import Workbook

from app import app
from prompts import SEARCH_MODES


def _post_financial_audit(workbook: Workbook, filename: str = "audit.xlsx"):
    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)
    client = app.test_client()
    return client.post(
        "/api/audit/financial",
        data={"file": (payload, filename), "analysis_type": "full"},
        content_type="multipart/form-data",
    )


def _high_signal_workbook() -> Workbook:
    wb = Workbook()
    low = wb.active
    low.title = "Kontrola"
    low["A1"] = "Kontrola"
    low["A2"] = '=IF(1=1,"OK","BŁĄD")'

    high = wb.create_sheet("Wysoki")
    high["A1"] = "Kwota"
    high["A2"] = "=10"
    high["A3"] = 15
    high["A4"] = "=20"
    return wb


def _hidden_reference_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jawny"
    ws["A1"] = "Opis"
    ws["B1"] = "Wynik"
    ws["A2"] = "rekord kontrolny"
    hidden = wb.create_sheet("UkryteDane")
    hidden.sheet_state = "hidden"
    hidden["A2"] = 123
    ws["B2"] = "=UkryteDane!A2"
    return wb


def test_financial_audit_builds_ai_evidence_pack_only_for_high_priority_findings():
    response = _post_financial_audit(_high_signal_workbook(), "ai-pack.xlsx")
    data = response.get_json()

    evidence_pack = data["ai_evidence_pack"]
    serialized = json.dumps(evidence_pack, ensure_ascii=False)
    findings = [
        finding
        for sheet in evidence_pack["sheets"]
        for finding in sheet.get("findings", [])
    ]

    assert response.status_code == 200
    assert data["success"] is True
    assert evidence_pack["summary"]["allowed_severities"] == ["HIGH", "CRITICAL"]
    assert findings
    assert all(finding["severity"] in {"HIGH", "CRITICAL"} for finding in findings)
    assert all(finding["type"] != "cosmetic_check" for finding in findings)
    assert "technical_appendix" not in serialized
    assert "screening_findings" not in serialized


def test_financial_audit_includes_high_hidden_sheet_reference_in_ai_evidence_pack():
    response = _post_financial_audit(_hidden_reference_workbook(), "hidden-pack.xlsx")
    data = response.get_json()

    evidence_pack = data["ai_evidence_pack"]
    findings = [
        finding
        for sheet in evidence_pack["sheets"]
        for finding in sheet.get("findings", [])
    ]

    assert response.status_code == 200
    assert data["success"] is True
    assert any(finding["type"] == "cross_sheet_hidden_reference" for finding in findings)
    hidden_ref = next(finding for finding in findings if finding["type"] == "cross_sheet_hidden_reference")
    assert hidden_ref["severity"] == "HIGH"
    assert hidden_ref["confidence_basis"] == "high_hidden_cross_sheet_dependency"
    assert hidden_ref["fraud_hypothesis"]
    assert hidden_ref["verification_target"]


def test_financial_forensics_prompt_requires_json_only():
    mode = SEARCH_MODES["financial_forensics"]

    assert "JSON" in mode["system"]
    assert "wyłącznie JSON" in mode["system"]
    assert "bez bloku ```json" in mode["system"]
    assert "findings[]" in mode["prompt_suffix"]


@patch("routes.financial_routes.call_llm")
def test_financial_opinion_endpoint_rejects_missing_high_priority_pack(mock_call_llm):
    wb = Workbook()
    ws = wb.active
    ws.title = "Czysty"
    ws["A1"] = 1
    ws["B1"] = 2
    ws["C1"] = "=A1+B1"

    audit_response = _post_financial_audit(wb, "clean.xlsx")
    report = audit_response.get_json()
    client = app.test_client()

    response = client.post(
        "/api/audit/financial/opinion",
        json={"evidence_pack": report.get("ai_evidence_pack")},
    )
    data = response.get_json()

    assert response.status_code == 400
    assert data["success"] is False
    assert data["opinion_available"] is False
    mock_call_llm.assert_not_called()


@patch("routes.financial_routes.call_llm")
def test_financial_opinion_endpoint_handles_invalid_model_json_fail_closed(mock_call_llm):
    audit_response = _post_financial_audit(_high_signal_workbook(), "invalid-json.xlsx")
    report = audit_response.get_json()
    client = app.test_client()
    mock_call_llm.return_value = {"response": "to nie jest json"}

    response = client.post(
        "/api/audit/financial/opinion",
        json={"evidence_pack": report["ai_evidence_pack"]},
    )
    data = response.get_json()

    assert response.status_code == 200
    assert data["success"] is True
    assert data["opinion_available"] is False
    assert data["ai_forensic_opinion"] is None
    assert "JSON" in data["warning"]
    mock_call_llm.assert_called_once()


@patch("routes.financial_routes.call_llm")
def test_financial_opinion_endpoint_returns_normalized_llm_opinion(mock_call_llm):
    audit_response = _post_financial_audit(_high_signal_workbook(), "valid-json.xlsx")
    report = audit_response.get_json()
    client = app.test_client()
    mock_call_llm.return_value = {
        "response": json.dumps(
            {
                "overall_assessment": "Silna poszlaka ręcznego obejścia logiki.",
                "sheet_comment": "Najmocniejszy sygnał dotyczy arkusza Wysoki.",
                "findings": [
                    {
                        "finding_id": "Wysoki:Wysoki!A3:hardcoded_in_formula_region",
                        "fact": "W komórce A3 wpisano ręcznie wartość 15 między formułami.",
                        "intent": "Możliwe narzucenie wyniku korzystnego dla autora.",
                        "expert_comment": "To wygląda jak punktowe obejście wzorca liczenia.",
                        "confidence": "wysoka",
                        "next_check": "Przelicz kolumnę według wzorca z komórek A2 i A4.",
                    }
                ],
            },
            ensure_ascii=False,
        )
    }

    response = client.post(
        "/api/audit/financial/opinion",
        json={"evidence_pack": report["ai_evidence_pack"]},
    )
    data = response.get_json()

    assert response.status_code == 200
    assert data["success"] is True
    assert data["opinion_available"] is True
    assert data["ai_forensic_opinion"]["overall_assessment"]
    assert data["ai_forensic_opinion"]["findings"][0]["finding_id"]
    assert data["ai_forensic_opinion"]["findings"][0]["fact"]
    assert data["provider"] == "ollama"
