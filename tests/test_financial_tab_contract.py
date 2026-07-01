from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from app import app


def test_financial_endpoint_returns_success_shape():
    client = app.test_client()

    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["B1"] = 2
    ws["C1"] = "=A1+B1"
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)

    response = client.post(
        "/api/audit/financial",
        data={"file": (payload, "audit.xlsx"), "analysis_type": "full"},
        content_type="multipart/form-data",
    )

    data = response.get_json()
    assert response.status_code == 200
    assert data["success"] is True
    assert data["file_name"] == "audit.xlsx"
    assert isinstance(data["summary"]["anomalies"], list)
    assert "file_metadata" in data
    assert "logic_overview" in data
    assert "description" in data["logic_overview"]
    first_sheet = next(iter(data["sheets"].values()))
    assert "logic_summary" in first_sheet
    assert "description_pl" in first_sheet["logic_summary"]
    assert "evidence_summary" in first_sheet
    assert "items" in first_sheet["evidence_summary"]
    assert "flow_graph" in first_sheet
    assert "nodes" in first_sheet["flow_graph"]
    assert "edges" in first_sheet["flow_graph"]


def test_show_tab_registry_includes_financial():
    template = Path("templates/index.html").read_text(encoding="utf-8")
    script = Path("static/financial_audit.js").read_text(encoding="utf-8")

    assert "tabFinancial" in template
    assert "financial:14" in template
    assert "financialAnalysisMode" in template
    assert "financialFlowExplorer" in template
    assert "financialAiOpinion" in template
    assert "financialRunButton" in template
    assert "handleFinancialFileSelection" in template
    assert "runFinancialAnalysis" in script
    assert "financialIsHighPriority" in script
    assert "loadFinancialOpinion" in script
    assert "renderFinancialOpinion" in script


def test_financial_endpoint_detects_cosmetic_check():
    client = app.test_client()

    wb = Workbook()
    ws = wb.active
    ws.title = "Kontrola"
    ws["A2"] = 10
    ws["C2"] = '=IF(A2=A2,"OK","BŁĄD")'
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)

    response = client.post(
        "/api/audit/financial",
        data={"file": (payload, "kontrola.xlsx"), "analysis_type": "full"},
        content_type="multipart/form-data",
    )

    data = response.get_json()
    anomaly_types = [item["type"] for item in data["summary"]["anomalies"]]
    first_sheet = next(iter(data["sheets"].values()))
    assert response.status_code == 200
    assert data["success"] is True
    assert "cosmetic_check" in anomaly_types
    assert any(
        "Kontrola!C2" in item and "pozorna" in item
        for item in first_sheet["evidence_summary"]["items"]
    )


def test_financial_endpoint_quick_mode_stops_after_screening_for_clean_sheet():
    client = app.test_client()

    wb = Workbook()
    ws = wb.active
    ws.title = "Czysty"
    ws["A1"] = 1
    ws["B1"] = 2
    ws["C1"] = "=A1+B1"
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)

    response = client.post(
        "/api/audit/financial",
        data={"file": (payload, "clean.xlsx"), "analysis_type": "quick"},
        content_type="multipart/form-data",
    )

    data = response.get_json()
    first_sheet = next(iter(data["sheets"].values()))
    assert response.status_code == 200
    assert data["success"] is True
    assert data["analysis"]["requested_mode"] == "quick"
    assert data["analysis"]["deep_scan_sheets"] == 0
    assert first_sheet["analysis"]["mode"] == "screening_only"
    assert first_sheet["analysis"]["deep_scan_triggered"] is False
    assert first_sheet["flow_graph"]["high_risk_relation_count"] == 0
    assert first_sheet["flow_graph"]["nodes"] == []


def test_financial_endpoint_targeted_mode_escalates_around_suspicious_area():
    client = app.test_client()

    wb = Workbook()
    ws = wb.active
    ws.title = "Trop"
    ws["A1"] = "Wartość"
    ws["B1"] = "Kontrola"
    ws["A2"] = 10
    ws["A3"] = 12
    ws["B2"] = '=IF(A2=A2,"OK","BŁĄD")'
    ws["B3"] = '=IF(A3=A3,"OK","BŁĄD")'
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)

    response = client.post(
        "/api/audit/financial",
        data={"file": (payload, "trace.xlsx"), "analysis_type": "targeted"},
        content_type="multipart/form-data",
    )

    data = response.get_json()
    first_sheet = next(iter(data["sheets"].values()))
    anomaly_types = [item["type"] for item in data["summary"]["anomalies"]]
    assert response.status_code == 200
    assert data["success"] is True
    assert data["analysis"]["requested_mode"] == "targeted"
    assert data["analysis"]["deep_scan_sheets"] == 1
    assert "cosmetic_check" in anomaly_types
    assert first_sheet["analysis"]["mode"] == "targeted_deep_scan"
    assert first_sheet["analysis"]["deep_scan_triggered"] is True
    assert "B" in first_sheet["analysis"]["focus_columns"]
    assert first_sheet["flow_graph"]["high_risk_relation_count"] == 0
    assert first_sheet["flow_graph"]["nodes"] == []


def test_financial_endpoint_reuses_screening_when_upgrading_mode_for_same_file():
    client = app.test_client()

    wb = Workbook()
    ws = wb.active
    ws.title = "ReuseFlow"
    ws["A1"] = "Wartość"
    ws["B1"] = "Kontrola"
    ws["A2"] = 101
    ws["A3"] = 102
    ws["B2"] = '=IF(A2=A2,"OK","BŁĄD")'
    ws["B3"] = '=IF(A3=A3,"OK","BŁĄD")'
    payload = BytesIO()
    wb.save(payload)
    raw = payload.getvalue()

    quick_response = client.post(
        "/api/audit/financial",
        data={"file": (BytesIO(raw), "reuse.xlsx"), "analysis_type": "quick"},
        content_type="multipart/form-data",
    )
    assert quick_response.status_code == 200

    full_response = client.post(
        "/api/audit/financial",
        data={"file": (BytesIO(raw), "reuse.xlsx"), "analysis_type": "full"},
        content_type="multipart/form-data",
    )

    data = full_response.get_json()
    assert full_response.status_code == 200
    assert data["success"] is True
    assert data["analysis"]["requested_mode"] == "full"
    assert data["analysis"]["screening_reused"] is True


def test_financial_endpoint_detects_benford_and_outlier_signals():
    client = app.test_client()

    wb = Workbook()
    ws = wb.active
    ws.title = "Statystyka"
    ws["A1"] = "Kwota"
    for idx in range(2, 31):
        ws[f"A{idx}"] = 900 + idx
    ws["A31"] = 999999
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)

    response = client.post(
        "/api/audit/financial",
        data={"file": (payload, "stats.xlsx"), "analysis_type": "full"},
        content_type="multipart/form-data",
    )

    data = response.get_json()
    anomaly_types = [item["type"] for item in data["summary"]["anomalies"]]
    first_sheet = next(iter(data["sheets"].values()))
    assert response.status_code == 200
    assert data["success"] is True
    assert "benford_deviation" in anomaly_types
    assert "amount_outlier" in anomaly_types
    assert first_sheet["forensic_signals"]["data_signals"]["benford_deviations"]
    assert first_sheet["forensic_signals"]["data_signals"]["amount_outliers"]
    assert first_sheet["flow_graph"]["high_risk_relation_count"] == 0


def test_financial_endpoint_detects_hidden_sheet_reference_without_low_priority_temporal_noise():
    client = app.test_client()

    wb = Workbook()
    ws = wb.active
    ws.title = "Jawny"
    ws["A1"] = "Opis"
    ws["B1"] = "Wynik"
    ws["A2"] = "rekord kontrolny"
    hidden = wb.create_sheet("UkryteDane")
    hidden.sheet_state = "hidden"
    hidden["A2"] = 123
    ws["B2"] = "=UkryteDane!A2"
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)

    response = client.post(
        "/api/audit/financial",
        data={"file": (payload, "hidden_ref.xlsx"), "analysis_type": "full"},
        content_type="multipart/form-data",
    )

    data = response.get_json()
    anomaly_types = [item["type"] for item in data["summary"]["anomalies"]]
    first_sheet = data["sheets"]["Jawny"]
    assert response.status_code == 200
    assert data["success"] is True
    assert "cross_sheet_hidden_reference" in anomaly_types
    hidden_reference = next(
        item for item in data["summary"]["anomalies"] if item["type"] == "cross_sheet_hidden_reference"
    )
    assert "weekend_activity" not in anomaly_types
    assert "night_activity" not in anomaly_types
    assert hidden_reference["severity"] == "HIGH"
    assert first_sheet["forensic_signals"]["control_signals"]["cross_sheet_hidden_references"]
    assert "weekend_activity" not in first_sheet["forensic_signals"]["data_signals"]
    assert "night_activity" not in first_sheet["forensic_signals"]["data_signals"]
    assert "numbering_gaps" not in first_sheet["forensic_signals"]["data_signals"]
    assert "near_thresholds" not in first_sheet["forensic_signals"]["data_signals"]
    assert "round_amounts" not in first_sheet["forensic_signals"]["data_signals"]
    assert set(first_sheet["forensic_signals"]["data_signals"].keys()) == {
        "duplicates",
        "amount_outliers",
        "benford_deviations",
    }
    assert first_sheet["flow_graph"]["high_risk_relation_count"] >= 1
    assert first_sheet["flow_graph"]["nodes"]


def test_financial_endpoint_flow_graph_shows_only_high_risk_relations():
    client = app.test_client()

    wb = Workbook()
    ws = wb.active
    ws.title = "Wysokie"
    ws["A1"] = "Kwota"
    ws["A2"] = "=10"
    ws["A3"] = 15
    ws["A4"] = "=20"
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)

    response = client.post(
        "/api/audit/financial",
        data={"file": (payload, "high-risk-flow.xlsx"), "analysis_type": "full"},
        content_type="multipart/form-data",
    )

    data = response.get_json()
    first_sheet = data["sheets"]["Wysokie"]
    anomaly_types = [item["type"] for item in data["summary"]["anomalies"]]
    assert response.status_code == 200
    assert data["success"] is True
    assert "hardcoded_in_formula_region" in anomaly_types
    assert first_sheet["flow_graph"]["high_risk_relation_count"] >= 1
    assert first_sheet["flow_graph"]["nodes"]


def test_high_priority_sheet_filter_matches_report_intent():
    client = app.test_client()

    wb = Workbook()
    ws = wb.active
    ws.title = "Niski"
    ws["A1"] = "Kontrola"
    ws["A2"] = '=IF(1=1,"OK","BŁĄD")'

    high = wb.create_sheet("Wysoki")
    high["A1"] = "Kwota"
    high["A2"] = "=10"
    high["A3"] = 15
    high["A4"] = "=20"

    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)

    response = client.post(
        "/api/audit/financial",
        data={"file": (payload, "priority-filter.xlsx"), "analysis_type": "full"},
        content_type="multipart/form-data",
    )

    data = response.get_json()
    high_priority_sheets = [
        name for name, sheet in data["sheets"].items()
        if sheet["risk_level"] in {"HIGH", "CRITICAL"}
    ]
    assert response.status_code == 200
    assert data["success"] is True
    assert "Wysoki" in high_priority_sheets
    assert "Niski" not in high_priority_sheets
